import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import load_workbook

import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = r"c:\Users\Raphael\Downloads\Allan FCL - Moneyball FM26 (1)\1. Planilha - Moneyball\Moneyball FM26 - Avancados.xlsm"

try:
    wb = load_workbook(file_path, data_only=True)
    # Check sheets
    target_sheet = [s for s in wb.sheetnames if "avancados" in s.lower() or "base" in s.lower() or "moneyball" in s.lower()]
    sheet = wb[target_sheet[0]] if target_sheet else wb.active
    
    # We need the base sheet or wherever data is pasted
    # Let's find the header row
    header_row = 1
    for r in range(1, 10):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 40)]
        if any(v and "Jogador" in str(v) for v in row_vals):
            header_row = r
            break
            
    from openpyxl.utils import get_column_letter

    headers = {}
    for c in range(1, 400):
        val = sheet.cell(row=header_row, column=c).value
        # also check the row above because sometimes headers are merged
        val2 = sheet.cell(row=header_row-1, column=c).value if header_row > 1 else None
        
        name = str(val).strip() if val else str(val2).strip() if val2 else ""
        if name and name != "None":
            col_letter = get_column_letter(c)
            headers[col_letter] = name

    out_lines = []
    for k, v in headers.items():
        out_lines.append(f"{k}: {v}")

    with open("excel_headers_mapping.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))
        
    print("Extracted headers to excel_headers_mapping.txt")

except Exception as e:
    print("Error:", e)
