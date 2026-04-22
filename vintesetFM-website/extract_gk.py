import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('e:\\fm26-editor-workspace-main\\fm26-editor-workspace\\vintesetFM-website\\Moneyball FM26 - Goleiros.xlsm', data_only=False)
ws = wb['🧤Goleiros']

headers = [str(cell.value) for cell in ws[1]]
row2 = [str(cell.value) for cell in ws[2]]

for h, r in zip(headers, row2):
    print(f"{h} ||| {r}")
