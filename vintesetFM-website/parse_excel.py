import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('c:/Users/Raphael/Downloads/Allan FCL - Moneyball FM26 (1)/1. Planilha - Moneyball/Moneyball FM26 - Avancados.xlsm', data_only=False)
sheet = wb.active
print(f"Sheet name: {sheet.title}")

# Let's inspect column headers and row 2 for formulas and conditional formatting
for col in range(1, min(sheet.max_column, 20)):
    top_cell = sheet.cell(row=1, column=col)
    second_cell = sheet.cell(row=2, column=col)
    
    print(col, f"Header: {top_cell.value}")
    if second_cell.data_type == 'f':
        print(f"  Formula in row 2: {second_cell.value}")
        
if sheet.conditional_formatting:
    print("\nConditional Formatting rules found:")
    for rule_range, rules in sheet.conditional_formatting._cf_rules.items():
        print(f"Range: {rule_range}")
        for rule in rules:
            color_scale_str = "yes" if rule.colorScale else "no"
            print(f" Rule: type={rule.type}, formulas={rule.formula}, colorScale={color_scale_str}")
