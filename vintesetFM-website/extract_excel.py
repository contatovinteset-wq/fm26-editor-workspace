# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import json

wb = load_workbook('c:/Users/Raphael/Downloads/Allan FCL - Moneyball FM26 (1)/1. Planilha - Moneyball/Moneyball FM26 - Avancados.xlsm', data_only=False)
sheet = wb.active

data = {"columns": [], "formulas": {}, "cf_rules": {}}

for col in range(1, sheet.max_column + 1):
    top_cell = sheet.cell(row=1, column=col)
    second_cell = sheet.cell(row=2, column=col)
    col_letter = top_cell.column_letter
    col_name = str(top_cell.value)
    
    data["columns"].append({"col": col_letter, "index": col, "name": col_name})
    
    if second_cell.data_type == 'f':
        data["formulas"][col_letter] = {"name": col_name, "formula": second_cell.value}

if sheet.conditional_formatting:
    for rule_range, rules in sheet.conditional_formatting._cf_rules.items():
        rule_list = []
        for rule in rules:
            r = {"type": rule.type}
            if rule.colorScale: r["colorScale"] = True
            if rule.dataBar: r["dataBar"] = True
            if rule.iconSet: r["iconSet"] = True
            rule_list.append(r)
        data["cf_rules"][str(rule_range)] = rule_list

with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("Done")
