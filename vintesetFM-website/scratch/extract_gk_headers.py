import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(
    r'c:\Users\Raphael\Downloads\Allan FCL - Moneyball FM26 (1)\1. Planilha - Moneyball\Moneyball FM26 - Goleiros.xlsm',
    data_only=True
)

ws = None
for sheet_name in wb.sheetnames:
    if 'Dado' in sheet_name or 'Data' in sheet_name or 'dado' in sheet_name:
        ws = wb[sheet_name]
        break
if ws is None:
    ws = wb[wb.sheetnames[0]]

print(f"Sheet: {repr(ws.title)}")
print(f"\n=== ROW 1 HEADERS (ALL) ===")
for col in range(1, min(ws.max_column + 1, 100)):
    val = ws.cell(row=1, column=col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"  {col_letter} (col {col}): {repr(val)}")
