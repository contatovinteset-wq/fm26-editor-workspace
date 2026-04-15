import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import load_workbook

# Change console encoding to utf-8 just in case
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = r"c:\Users\Raphael\Downloads\Allan FCL - Moneyball FM26 (1)\1. Planilha - Moneyball\Moneyball FM26 - Avancados.xlsm"

try:
    wb = load_workbook(file_path, data_only=False)
    # Find the moneyball sheet
    target_sheet = [s for s in wb.sheetnames if "moneyball" in s.lower() or "avançados" in s.lower() or "avancado" in s.lower() or "base" in s.lower()]
    sheet = wb[target_sheet[0]] if target_sheet else wb.active
    
    header_row = 1
    for r in range(1, 10):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 40)]
        if any(v and "Jogador" in str(v) for v in row_vals):
            header_row = r
            break
            
    data_row = header_row + 1
    formulas = {}
    
    out_lines = []
    
    for c in range(1, 150):
        header = sheet.cell(row=header_row, column=c).value
        val = sheet.cell(row=data_row, column=c).value
        # extract conditional formatting rules for this column? that's harder in openpyxl, but we can do it if needed
        # for now let's just get the formula
        if header and str(header).strip():
            formulas[str(header).strip()] = {
                'col': c,
                'formula': str(val) if str(val).startswith('=') else f"Data (e.g. {val})",
            }
            
    for k, v in formulas.items():
        out_lines.append(f"{k}: {v['formula']}")

    with open("excel_formulas.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(out_lines))
        
    print("Extracted formulas to excel_formulas.txt")

except Exception as e:
    print("Error:", e)
