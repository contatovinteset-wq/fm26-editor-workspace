import openpyxl
import json

file_path = r"c:\Users\Raphael\Downloads\Allan FCL - Moneyball FM26 (1)\1. Planilha - Moneyball\Moneyball FM26.xlsm"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    all_logic = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        headers = []
        formulas = []
        
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=False)):
            if idx == 0:
                headers = [cell.value for cell in row]
            elif idx == 4:
                formulas = [cell.value for cell in row]
                break
                
        header_formula_map = []
        for h, f in zip(headers, formulas):
            if h and str(h).strip() != '':
                header_formula_map.append({"Header": str(h), "Formula/Value": str(f)})
                
        all_logic[sheet_name] = header_formula_map
        
    with open("all_moneyball_logic.json", "w", encoding="utf-8") as file:
        json.dump(all_logic, file, indent=2, ensure_ascii=False)
        
    print("Logic dumped to all_moneyball_logic.json")
    
except Exception as e:
    import traceback
    traceback.print_exc()
